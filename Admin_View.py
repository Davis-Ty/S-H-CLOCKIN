
from datetime import date
from tkinter import Button, Entry, Label, PhotoImage, Tk
import threading
import os
import openpyxl 
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import xlsxwriter


#(Functiom background_pic) take pic and sets it to the background
def background_pic(pic):
    background=Label(root,image=pic,)
    background.place(x=100,y=0)

    


def adminPass():
    if (ea.get()!="Admin15" or eb.get()!="Teacher!15"):
        returns ="Wrong Username or Password!"
        userLabel= Label(root, text=returns) 
        userLabel.pack()
        threading.Timer(1.0,userLabel.pack_forget).start()
        ea.delete(0,"end")  
        eb.delete(0,"end")
        ea.insert(0,"Username: ",)
        ea.bind("<FocusIn>",temp_texta)
        eb.insert(0,"Password: ",)
        eb.bind("<FocusIn>",temp_textb) 
        __name__
        
    else:
        
        #(Function) to get the total pay
        def calculate_total_pay(sheet, row):

            #(users) clock-In time
            num=float(((str(sheet[f'E{row}'].value)[:8]).replace(":","")))
            formatted_num = float('{:.4f}'.format(num/10000))

            #(users) clock-Out time
            num2=float(((str(sheet[f'F{row}'].value)[:8]).replace(":","")))
            formatted_num2 = float('{:.4f}'.format(num2/10000))

            #(users) lunch-In time
            num3=float(((str(sheet[f'G{row}'].value)[:8]).replace(":","")))
            formatted_num3 = float('{:.4f}'.format(num3/10000))

            #(users) lunch-Out time
            num4=float(((str(sheet[f'H{row}'].value)[:8]).replace(":","")))
            formatted_num4 = float('{:.4f}'.format(num4/10000))

            #(getting total num of hrs)
            hours_worked = ((formatted_num2-formatted_num)-(formatted_num4-formatted_num3))
            
            #getting pay amount by hr
            hourly_rate = sheet[f'I{row}'].value


            formula = float(hours_worked)*float(hourly_rate)
            
            #returning fomula (PAY)
            return formula


        #(Function)checking if value is None 
        def is_valid_id(sheet, row):
            return sheet['B' + str(row)].value is not None

        # (Function) opens file that is already made calls caulate function stores data in array prints it onto file(.txt)
        def payRoll():
            try:
                todayDate = date.today()
                numDayOfWeek=int(todayDate.strftime("%d"))
                week=1
                
                for i in range (1,numDayOfWeek+1):
                    check= todayDate.replace(day=i)
                    checkday=check.strftime("%A")
                    if (checkday=='Sunday'):
                        week =week+1

                filename = "payroll" + todayDate.strftime("%B%Y") + "Week" + str(week) + ".txt"

                ans=0.0000
                DayOfWeek=["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
                todayDate = date.today()
                numDayOfWeek=int(todayDate.strftime("%d"))
                write=[]
            
                # Load the workbook and select the sheet
                wb = openpyxl.load_workbook("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
                for row in range(2, 200):
                    for sheet_name in DayOfWeek:
                        sheet = wb[sheet_name]    
                        if is_valid_id(sheet, row):
                            total_pay = calculate_total_pay(sheet, row)
                            employee_id = sheet['B' + str(row)].value
                            FirstName=sheet['C' + str(row)].value
                            LastName=sheet['D' + str(row)].value
                            ans=ans+total_pay
                        else:
                            break  # no more employees on this sheet
                    else:
                        if (round(ans,2)>0):
                            write.append(f"{employee_id} {FirstName} {LastName} Total Pay: {round(ans,2)}")
                            ans=0.0000
                            
                            
                        elif (round(ans,2)==0):
                            write.append(f"{employee_id} {FirstName} {LastName} DID NOT WORK THIS WEEK!! Total Pay: {round(ans,2)}")
                            ans=0.0000
                            
                        else: 
                            write.append(f"{employee_id} {FirstName} {LastName} FORGOT TO CLOCK-OUT!! Total Pay: {round(ans,2)}")
                            ans=0.0000
                            start=start+1
                
                
                    
                with open(filename, "w") as f:
                    for ppl in write:
                        f.write(str(ppl) + "\n")
                    
                    returns="Look for file named payroll" + todayDate.strftime("%B%Y") + "Week" + str(week) + ".txt"
                    userLabel= Label(root, text=returns) 
                    userLabel.pack()
                    threading.Timer(5.0,userLabel.pack_forget).start()
            except: 
                returns="File ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx is unavailable"
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(5.0,userLabel.pack_forget).start()
            

        def exeDatabase():
            if (os.path.isfile("ShinningHillEmployee.xlsx")==False): 
                #name of file
                    workbook=xlsxwriter.Workbook("ShinningHillEmployee.xlsx")
                    
                    worksheet=workbook.add_worksheet("workers")
                    #input the data to file
                    for i in range(0,7):
                        

                        worksheet.write(0,0,"Date")
                        worksheet.write(0,1,"Employee ID")
                        worksheet.write(0,2,"First Name")
                        worksheet.write(0,3,"Last Name")
                        worksheet.write(0,4,"Clock in time")
                        worksheet.write(0,5,"Clock out time")
                        worksheet.write(0,6,"Lunch Break Start Time")
                        worksheet.write(0,7,"Lunch Break End Time")
                        worksheet.write(0,8,"Pay Rate")
                        worksheet.write(0,9,"Total Pay")
                        worksheet.write(0,10,"Total Time Worked")
                    workbook.close() 

        def loadup():
            try:
                iD=int(e.get())
                fName=str(e1.get())
                lName= str(e2.get())
                pay=(e3.get())
                ID="B"
                fn="C"
                ln="D"
                ci="E"
                co="F"
                li="G"
                lo="H"
                p="I"
                tp="J"
                tt="K"
                i=1
                    
                wb= openpyxl.load_workbook("ShinningHillEmployee.xlsx")

                #gives us the sheet we want (day of week)
                wb.active=wb["workers"]
                sheet=wb.active
                        
                        

                while not sheet[ID+str(i)].value is None:
                        if (sheet[ID+str(i)].value==iD):
                            returns="ID is taken"
                            userLabel= Label(root, text=returns) 
                            userLabel.pack()
                            threading.Timer(1.0,userLabel.pack_forget).start()
                            e.delete(0,"end")
                            e1.delete(0,"end")
                            e2.delete(0,"end")
                            e3.delete(0,"end")
                            e.insert(0,"User ID#: ",)
                            e.bind("<FocusIn>",temp_text)
                            e1.insert(0,"User ID#: ",)
                            e1.bind("<FocusIn>",temp_text1)
                            e2.insert(0,"User ID#: ",)
                            e2.bind("<FocusIn>",temp_text2)
                            e3.insert(0,"User ID#: ",)
                            e3.bind("<FocusIn>",temp_text3) 
                            
                            exit()
                            
                            
                        else:i=i+1
                else:
                        sheet[ID+str(i)].value= iD
                        sheet[fn+str(i)].value=fName
                        sheet[ln+str(i)].value=lName
                        sheet[ci+str(i)].value=0
                        sheet[co+str(i)].value=0
                        sheet[li+str(i)].value=0
                        sheet[lo+str(i)].value=0
                        sheet[p+str(i)].value= str(pay)
                        sheet[tp+str(i)].value="=(K"+str(i)+"*I"+str(i)+")"
                        sheet[tt+str(i)].value="=((F"+str(i)+"-E"+str(i)+")+(H"+str(i)+"-G"+str(i)+"))*24"
                        wb.save("ShinningHillEmployee.xlsx")
            
                returns="User was added"
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(1.0,userLabel.pack_forget).start()
                e.delete(0,"end")
                e1.delete(0,"end")
                e2.delete(0,"end")
                e3.delete(0,"end")
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)
                e1.insert(0,"First Name",)
                e1.bind("<FocusIn>",temp_text1)
                e2.insert(0,"Last Name",)
                e2.bind("<FocusIn>",temp_text2)
                e3.insert(0,"Pay: ",)
                e3.bind("<FocusIn>",temp_text3)  
            except:        
                returns="user WAS NOT ADDED (ID MAY BE TAKEN)"
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(1.0,userLabel.pack_forget).start()
                e.delete(0,"end")
                e1.delete(0,"end")
                e2.delete(0,"end")
                e3.delete(0,"end")
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)
                e1.insert(0,"First Name: ",)
                e1.bind("<FocusIn>",temp_text1)
                e2.insert(0,"Last Name: ",)
                e2.bind("<FocusIn>",temp_text2)
                e3.insert(0,"Pay: ",)
                e3.bind("<FocusIn>",temp_text3)
            

        def delete():
            try:
                iD=int(e.get())
                ID="B"
                fn="C"
                ln="D"
                ci="E"
                co="F"
                li="G"
                lo="H"
                p="I"
                tp="J"
                tt="K"
                i=1
                wb= openpyxl.load_workbook("ShinningHillEmployee.xlsx")

                #gives us the sheet we want (day of week)
                wb.active=wb["workers"]
                sheet=wb.active
                    
                    

                while not sheet[ID+str(i)].value==iD:
                    
                    i=i+1
                    if(i==101):
                        returns="Looking... "
                        userLabel= Label(root, text=returns) 
                        userLabel.pack()
                        threading.Timer(1.0,userLabel.pack_forget).start()
                        e.delete(0,"end")
                        e.insert(0,"User ID#: ",)
                        e.bind("<FocusIn>",temp_text)
                        exit()

                else:
                    sheet[ID+str(i)].value= None
                    sheet[fn+str(i)].value=None
                    sheet[ln+str(i)].value=None
                    sheet[ci+str(i)].value=None
                    sheet[co+str(i)].value=None
                    sheet[li+str(i)].value=None
                    sheet[lo+str(i)].value=None
                    sheet[p+str(i)].value= None
                    sheet[tp+str(i)].value=None
                    sheet[tt+str(i)].value=None
                    
                    wb.save("ShinningHillEmployee.xlsx")
                    
                returns="User was removed"
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(1.0,userLabel.pack_forget).start()
                e.delete(0,"end")
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)
                
                    
            except:
                returns="no ID found "
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(1.0,userLabel.pack_forget).start()
                e.delete(0,"end")
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)
                
                
        def replace():
            try:
                iD=int(e.get())
                rD=int(e4.get())
                
                ID="B"

                i=1
                wb= openpyxl.load_workbook("ShinningHillEmployee.xlsx")

                #gives us the sheet we want (day of week)
                wb.active=wb["workers"]
                sheet=wb.active
                
                    

                while not sheet[ID+str(i)].value==iD:
                    
                    i=i+1
                else:
                    sheet[ID+str(i)].value= rD
                    
                    wb.save("ShinningHillEmployee.xlsx")
                    returns="ID was replaced"
                    userLabel= Label(root, text=returns) 
                    userLabel.pack()
                    threading.Timer(1.0,userLabel.pack_forget).start()
                    e.delete(0,"end")
                    e4.delete(0,"end")
                    e.insert(0,"User ID#: ",)
                    e.bind("<FocusIn>",temp_text)
                    e4.insert(0,"ID REPLACEMENT: ",)
                    e4.bind("<FocusIn>",temp_text4) 
            except:
                returns ="no ID found "
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(1.0,userLabel.pack_forget).start()
                e.delete(0,"end")  
                e4.delete(0,"end")
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)
                e4.insert(0,"ID REPLACEMENT: ",)
                e4.bind("<FocusIn>",temp_text4) 


        def replacePay():
            try:
                iD=int(e.get())
                pay=(e5.get())
                
                ID="B"
                p="I"

                i=1
                wb= openpyxl.load_workbook("ShinningHillEmployee.xlsx")

                #gives us the sheet we want (day of week)
                wb.active=wb["workers"]
                sheet=wb.active
                
                    

                while not sheet[ID+str(i)].value==iD:
                    
                    i=i+1
                else:
                    sheet[p+str(i)].value= str(pay)
                    
                    wb.save("ShinningHillEmployee.xlsx")
                    returns="Pay was replaced"
                    userLabel= Label(root, text=returns) 
                    userLabel.pack()
                    threading.Timer(1.0,userLabel.pack_forget).start()
                    e.delete(0,"end")
                    e5.delete(0,"end")
                    e.insert(0,"User ID#: ",)
                    e.bind("<FocusIn>",temp_text)
                    e5.insert(0,"PAY REPLACEMENT: ",)
                    e5.bind("<FocusIn>",temp_text5) 
            except:
                returns ="Check user ID "
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(1.0,userLabel.pack_forget).start()
                e.delete(0,"end")  
                e5.delete(0,"end")
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)
                e5.insert(0,"PAY REPLACEMENT: ",)
                e5.bind("<FocusIn>",temp_text5) 


        def clock_change():
            try:
                todayDate = date.today()
                numDayOfWeek=int(todayDate.strftime("%d"))
                week=1
                for i in range (1,numDayOfWeek+1):
                    check= todayDate.replace(day=i)
                    checkday=check.strftime("%A")
                    if (checkday=='Sunday'):
                        week =week+1
                user=e.get()
                ID="B"
                clock_in="E"
                clock_out="F"
                Lunch_in="G"
                Lunch_out="H"
                returns=""
                Day=str(eday.get()).replace(" ","")
                #get month/week/day
                todayDate = date.today()
                if (os.path.isfile("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")==True):
                    print()
                    wb = openpyxl.load_workbook("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
                    #gives us the sheet we want (day of week)
                    wb.active=wb[Day]
                    sheet=wb.active
                    user=int(e.get())

                    #testig if ID is valid (Remember make it a function later)
                    i=1
                
                    while not (sheet[ID+str(i)].value==str(user)):
                        i=i+1
                        if(i==100):
                            returns ="ID not stored"
                    
                            break
                    else:
                        #lets user go to lunch if it passes the constrains 
                        if (eli.get() is not None):
                            sheet[Lunch_in+str(i)].value=str(eli.get()).replace("HH:MM:SS Lunch-In","00:00:00")
                            returns="Times have been changed"
                        if (elo.get() is not None):
                            sheet[Lunch_out+str(i)].value=str(elo.get()).replace("HH:MM:SS Lunch-Out","00:00:00")
                            returns="Times have been changed"
                        if (eci.get() is not None):
                            sheet[clock_in+str(i)].value=str(eci.get()).replace("HH:MM:SS Clock-In","00:00:00")
                            returns="Times have been changed"
                        if (eco.get() is not None):
                            sheet[clock_out+str(i)].value=str(eco.get()).replace("HH:MM:SS Clock-Out","00:00:00")
                            returns="Times have been changed"
                else:
                    returns= "No file found"
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(3.0,userLabel.pack_forget).start()

                e.delete(0,"end")  
                eday.delete(0,"end")
                
                eci.delete(0,"end")
                eco.delete(0,"end")
                eli.delete(0,"end")
                elo.delete(0,"end")
                        
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)


                eci.insert(0,"HH:MM:SS Clock-In",)
                eci.bind("<FocusIn>",temp_textci) 

                eco.insert(0,"HH:MM:SS Clock-Out",)
                eco.bind("<FocusIn>",temp_textco) 

                eli.insert(0,"HH:MM:SS Lunch-In",)
                eli.bind("<FocusIn>",temp_textli)

                elo.insert(0,"HH:MM:SS Lunch-Out",)
                elo.bind("<FocusIn>",temp_textlo) 
                eday.insert(0,"Monday",)
                eday.bind("<FocusIn>",temp_textday)

                wb.save("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
            except:
                returns= "ID not found"
                userLabel= Label(root, text=returns) 
                userLabel.pack()
                threading.Timer(3.0,userLabel.pack_forget).start()

                e.delete(0,"end")  
                eci.delete(0,"end")
                eco.delete(0,"end")
                eli.delete(0,"end")
                elo.delete(0,"end")
                        
                e.insert(0,"User ID#: ",)
                e.bind("<FocusIn>",temp_text)

                eci.insert(0,"HH:MM:SS Clock-In",)
                eci.bind("<FocusIn>",temp_textci) 

                eco.insert(0,"HH:MM:SS Clock-Out",)
                eco.bind("<FocusIn>",temp_textco) 

                eli.insert(0,"HH:MM:SS Lunch-In",)
                eli.bind("<FocusIn>",temp_textli)

                elo.insert(0,"HH:MM:SS Lunch-Out",)
                elo.bind("<FocusIn>",temp_textlo) 

        def temp_text(n):
            e.delete(0,"end")

        def temp_text1(n):
            e1.delete(0,"end")

        def temp_text2(n):
            e2.delete(0,"end")

        def temp_text3(n):
            e3.delete(0,"end")

        def temp_text4(n):
            e4.delete(0,"end")

        def temp_text5(n):
            e5.delete(0,"end")

        def temp_textli(n):
            eli.delete(0,"end")

        def temp_textlo(n):
            elo.delete(0,"end")

        def temp_textci(n):
            eci.delete(0,"end")

        def temp_textco(n):
            eco.delete(0,"end")

        def temp_textday(n):
            eday.delete(0,"end")

        def on_closing():
            root.protocol("WM_DELETE_WINDOW", disable_close)
            root.after(5000, root.destroy)

        def disable_close():
            pass
        
        exeDatabase()

        myButtona.pack_forget()
        ea.pack_forget()
        eb.pack_forget()

        
        #sets textbox features
        e=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        e.pack()
        #insert text into textbox
        e.insert(0,"User ID#: ",)

        #makes texted in box delete by calling the main temp_text function
        e.bind("<FocusIn>",temp_text)

        #sets textbox features
        e1=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        e1.pack()
        #insert text into textbox
        e1.insert(0,"First Name: ",)

        #makes texted in box delete by calling the main temp_text function
        e1.bind("<FocusIn>",temp_text1)


            #sets textbox features
        e2=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        e2.pack()
        #insert text into textbox
        e2.insert(0,"Last Name: ",)

        #makes texted in box delete by calling the main temp_text function
        e2.bind("<FocusIn>",temp_text2)


        #sets textbox features
        e3=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        e3.pack()
        #insert text into textbox
        e3.insert(0,"Pay:",)

        #makes texted in box delete by calling the main temp_text function
        e3.bind("<FocusIn>",temp_text3)

        e4=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        e4.pack()
        #insert text into textbox
        e4.insert(0,"ID replacement:",)

        #makes texted in box delete by calling the main temp_text function
        e4.bind("<FocusIn>",temp_text4)
        
        e5=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        e5.pack()
        #insert text into textbox
        e5.insert(0,"New Pay: ",)

        #makes texted in box delete by calling the main temp_text function
        e5.bind("<FocusIn>",temp_text5)

        eli=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        eli.pack()
        #insert text into textbox
        eli.insert(0,"HH:MM:SS Lunch-In",)

        #makes texted in box delete by calling the main temp_text function
        eli.bind("<FocusIn>",temp_textli)

        elo=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        elo.pack()
        #insert text into textbox
        elo.insert(0,"HH:MM:SS Lunch-Out",)

        #makes texted in box delete by calling the main temp_text function
        elo.bind("<FocusIn>",temp_textlo)

                
        eci=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        eci.pack()
        #insert text into textbox
        eci.insert(0,"HH:MM:SS Clock-In ",)

        #makes texted in box delete by calling the main temp_text function
        eci.bind("<FocusIn>",temp_textci)

        eco=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        eco.pack()
        #insert text into textbox
        eco.insert(0,"HH:MM:SS Clock-Out ",)

        #makes texted in box delete by calling the main temp_text function
        eco.bind("<FocusIn>",temp_textco)

        
        eday=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
        eday.pack()
        #insert text into textbox
        eday.insert(0,"Tuesday",)

        #makes texted in box delete by calling the main temp_text function
        eday.bind("<FocusIn>",temp_textday)
        

        
        
        #makes button return userClick function
        myButton= Button(root,text="ADD USER", command=loadup,padx=(10),pady=(10),background = "green", fg = "white")
        myButton.pack(padx=(5),pady=(5))
        

        myButton2= Button(root,text="REMOVE USER", command=delete,padx=(10),pady=(10),background = "red", fg = "white")

        myButton2.pack(padx=(5),pady=(5))

        myButton3= Button(root,text="CHANGE ID", command=replace,padx=(10),pady=(10),background = "green", fg = "white")

        myButton3.pack(padx=(5),pady=(5))

        myButton4= Button(root,text="NEW PAY", command=replacePay,padx=(10),pady=(10),background = "green", fg = "white")

        myButton4.pack(padx=(5),pady=(5))
        
        myButtonp= Button(root,text="Get Pay Roll", command=payRoll,padx=(10),pady=(10),background = "green", fg = "white")
        myButtonp.pack(padx=(5),pady=(5))

        myButtont= Button(root,text="Change Time", command=clock_change,padx=(10),pady=(10),background = "green", fg = "white")
        myButtont.pack(padx=(5),pady=(5))

        
        root.protocol("WM_DELETE_WINDOW", on_closing)
        root.mainloop()

def temp_texta(n):
        ea.delete(0,"end")

def temp_textb(n):
        eb.delete(0,"end")

if __name__ == '__main__':



    #setting tkinter
    root = Tk()


    #sets size of the gui that opens
    root.geometry("500x1000")


    photo_filename = "ShinningHillPic.png"
   
    bg=PhotoImage(file=photo_filename)#must pass pic to PhotoImage function and set it to variable
    #pass background pic to background_pic function
    background_pic(bg)


    #sets textbox features
    ea=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
    ea.pack()
    #insert text into textbox
    ea.insert(0,"User Name: ",)

    #makes texted in box delete by calling the main temp_text function
    ea.bind("<FocusIn>",temp_texta)

        #sets textbox features
    eb=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5,show="*")
    eb.pack()

    #makes texted in box delete by calling the main temp_text function
    eb.bind("<FocusIn>",temp_textb)


    
    
    #makes button return userClick function
    myButtona= Button(root,text="Login", command=adminPass,padx=(10),pady=(10),background = "green", fg = "white")
    myButtona.pack(padx=(10),pady=(10))
    


    
    root.mainloop()
    #(main line)-END 
