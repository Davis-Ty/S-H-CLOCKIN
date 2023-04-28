#@author Tyon Davis 


import threading
from tkinter import Button, Entry, Label, PhotoImage, Tk
from datetime import date
import datetime
import os
import openpyxl 
import time
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import xlsxwriter


def loadup():
    Date="A"
    ID="B"
    fn="C"
    ln="D"
    clock_in="E"
    clock_out="F"
    Lunch_in="G"
    Lunch_out="H"
    p="I"
    tp="J"
    tt="K"
    i=2

    data=[]
        
    wb1= openpyxl.load_workbook("ShinningHillEmployee.xlsx")

    #gives us the sheet we want (day of week)
    wb1.active=wb1["workers"]
    sheet=wb1.active
    while sheet[ID+str(i)].value is not None or sheet[ID+str(i+1)].value is not None or sheet[ID+str(i+2)].value is not None or sheet[ID+str(i+3)].value is not None:  

            #(employees data hardwired in)
                employee={
                        'Employee ID': str(sheet[ID+str(i)].value),
                        'First Name':str(sheet[fn+str(i)].value),
                        'Last Name':str(sheet[ln+str(i)].value),
                        'Clock in time':str(sheet[clock_in+str(i)].value),
                        'Clock out time':str(sheet[clock_out+str(i)].value),
                        'Lunch Break Start Time':str(sheet[Lunch_in+str(i)].value),
                        'Lunch Break End Time':str(sheet[Lunch_out+str(i)].value),
                        'Pay Rate':str(sheet[p+str(i)].value),
                        'Total Pay':"ASK ADMIN",
                        'Total Time Worked':str(sheet[tt+str(i)].value),
                }
                data.append(employee)
                i=i+1
        

    #List of days for data in each sheet of workbook
    DayOfWeek=["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]

    #date information to make files with new name
    todayDate = date.today()

    numDayOfWeek=int(todayDate.strftime("%d"))

    # finding what number week we are on sunday starts a week
    week=1
    for i in range (1,numDayOfWeek+1):
        check= todayDate.replace(day=i)
        checkday=check.strftime("%A")
        if (checkday=='Sunday'):
            week =week+1

    #checks if file was made already... if not makes a new file
    if (os.path.isfile("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")==True):
        print()
    else:

        #name of file
        workbook=xlsxwriter.Workbook("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
        i=0
        #input the data to file
        for i in range(0,7):
            worksheet=workbook.add_worksheet(DayOfWeek[i])

            worksheet.write(0,0,"Date")
            worksheet.write(0,1,"Employee ID")
            worksheet.write(0,2,"First Name")
            worksheet.write(0,3,"Last Name")
            worksheet.write(0,4,"Clock in time")
            worksheet.write(0,5,"Clock out time")
            worksheet.write(0,6,"Lunch Break Start Time")
            worksheet.write(0,7,"Lunch Break End Time")
            worksheet.write(0,8,"Pay Rate")
            worksheet.write(0,9,"Total Pay")
            worksheet.write(0,10,"Total Time Worked")
            

            for index, entry in enumerate(data):
                worksheet.write(index+1,0,str(" "))
                worksheet.write(index+1,1,entry["Employee ID"])
                worksheet.write(index+1,2,entry["First Name"])
                worksheet.write(index+1,3,entry["Last Name"])
                worksheet.write(index+1,4,entry["Clock in time"])
                worksheet.write(index+1,5,entry["Clock out time"])
                worksheet.write(index+1,6,entry["Lunch Break Start Time"])
                worksheet.write(index+1,7,entry["Lunch Break End Time"])
                worksheet.write(index+1,8,entry["Pay Rate"])
                worksheet.write(index+1,9,entry["Total Pay"])
                worksheet.write(index+1,10,entry["Total Time Worked"]) 
                 
        print()
        workbook.close()
        

#if user presses clockIn
def clockIn():
    try: 
        user=int(e.get())
        #testig if ID is valid (Remember make it a function later)
        i=1
    
        while not (sheet[ID+str(i)].value==str(user)):
            i=i+1
            if(i==100):
                returns ="ID not stored"
                break
        else:
            #clockng user in if user has not already clocked in
            sheet[Date+str(i)].value=todayDate.strftime("%B/%d/%Y")

            if (sheet[clock_in+str(i)].value=="0"):
                sheet[clock_in+str(i)].value=datetime.datetime.now().time()
                        
                returns= user ,"Clock In"
            else:returns="You have already clocked in!!"
                
                #save work
            wb.save("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
            #display output for a sec
            userLabel= Label(root, text=returns) 
            userLabel.pack()
            user=int(e.get())
            threading.Timer(1.0,userLabel.pack_forget).start()
            e.delete(0,"end")

    except:returns="ID needed!!!"
    userLabel= Label(root, text=returns) 
    userLabel.pack()
    threading.Timer(1.0,userLabel.pack_forget).start()
    e.delete(0,"end")


def clockOut():
    try:
        user=int(e.get())
        #testig if ID is valid (Remember make it a function later)
        i=1
    
        while not (sheet[ID+str(i)].value==str(user)):
            i=i+1
            if(i==100):
                returns ="ID not stored"
                break
        else:
            #lets user clock if it passes the constrains 
        
            if (sheet[clock_in+str(i)].value=="0"):
                returns ="you are not clocked In."
            
            else:
                #lets user clock out if it passes the constrains 
                if(sheet[clock_out+str(i)].value=="0"):
                        if(sheet[Lunch_in+str(i)].value!="0"):
                            if(sheet[Lunch_out+str(i)].value=="0"):
                                returns="you are still at lunch"
                    
                            else:   
                                returns=user,"clocked out"
                                sheet[clock_out+str(i)].value=datetime.datetime.now().time()
                                
                        else:   
                            returns=user,"clocked out"
                            sheet[clock_out+str(i)].value=datetime.datetime.now().time()
                else:returns="you have already clocked out!" 
        
        #save work
        wb.save("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
             
        #display output for a sec
        userLabel= Label(root, text=returns) 
        userLabel.pack()
        user=int(e.get())
        threading.Timer(1.0,userLabel.pack_forget).start()
        e.delete(0,"end")
    
    except:returns="ID needed!!!"
    userLabel= Label(root, text=returns) 
    userLabel.pack()
    threading.Timer(1.0,userLabel.pack_forget).start()
    e.delete(0,"end")


def lunchIn():
    try:
        user=int(e.get())

        #testig if ID is valid (Remember make it a function later)
        i=1
    
        while not (sheet[ID+str(i)].value==str(user)):
            i=i+1
            if(i==100):
                returns ="ID not stored"
                #user=int(e.get())
                break
        else:
            #lets user go to lunch if it passes the constrains 
         
            if (sheet[Lunch_in+str(i)].value=="0"):
                    if (sheet[clock_in+str(i)].value=="0"):
                        returns="you must clock in!"
                    else:
                        if(sheet[clock_out+str(i)].value!="0"):
                            returns="you are already clocked out!"
                        else:
                            sheet[Lunch_in+str(i)].value=datetime.datetime.now().time()
            
                            returns=user, "@ lunch"

        #save work
        wb.save("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")

        #display output for a sec
        userLabel= Label(root, text=returns) 
        userLabel.pack()
        user=int(e.get())
        threading.Timer(1.0,userLabel.pack_forget).start()
        e.delete(0,"end")

    except:returns="ID needed!!!"
    userLabel= Label(root, text=returns) 
    userLabel.pack()
    threading.Timer(1.0,userLabel.pack_forget).start()
    e.delete(0,"end")

def lunchOut():
    try:
        user=int(e.get())

        #testig if ID is valid (Remember make it a function later)
        i=1
        while not (sheet[ID+str(i)].value==str(user)):
            i=i+1
            if(i==100):
                returns ="ID not stored"
                break
        else:
            #lets user come back from lunch if it passes the constrains 
            if ( sheet[clock_in+str(i)].value!="0"):
                if ( sheet[Lunch_in+str(i)].value!="0"):
                    if ( sheet[Lunch_out+str(i)].value=="0"):
                        returns=user, "welcome back"
                        sheet[Lunch_out+str(i)].value=datetime.datetime.now().time()

                    else: 
                        returns="You already went to lunch!"
                else: returns= "You not clocked in for lunch!"
            else: returns="You have not clocked in!"
    
        #save work
        wb.save("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
        userLabel= Label(root, text=returns) 
        userLabel.pack()
        user=int(e.get())
        threading.Timer(1.0,userLabel.pack_forget).start()
        e.delete(0,"end")

    except:returns="ID needed!!!"
    userLabel= Label(root, text=returns) 
    userLabel.pack()
    threading.Timer(1.0,userLabel.pack_forget).start()
    e.delete(0,"end")

def temp_text(n):
    e.delete(0,"end")


#(Functiom background_pic) take pic and sets it to the background
def background_pic(pic):
    background=Label(root,image=pic,)
    background.place(x=100,y=0)


#(main line)-START

if __name__ == '__main__':
    loadup()
    
    todayDate = date.today()
    
    week =5


    
    #searches for file made by ShinningHillEmployeePopulate
    while not (os.path.isfile("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")):
        if(week>=0):
            week=week-1
        else:__name__
    
    #loads the workbook
    wb= openpyxl.load_workbook("ShinningHillEmployeeTimeCard"+todayDate.strftime("%B%Y")+"Week"+str(week)+".xlsx")
    

    #gives us the sheet we want (day of week)
    wb.active=wb[todayDate.strftime("%A")]
    sheet=wb.active

    #setting variables for data in excel file
    todayDate = date.today()
   

    Date="A"
    ID="B"
    fn="C"
    ln="D"
    clock_in="E"
    clock_out="F"
    Lunch_in="G"
    Lunch_out="H"
    p="I"
    tp="J"
    tt="K"
    pd="L"

    milTime = time.strftime("%H:%M:%S")

    #setting tkinter
    root = Tk()

    #sets size of the gui that opens
    root.geometry("500x1000")


    photo_filename = "ShinningHillPic.png"
   
    bg=PhotoImage(file=photo_filename)#must pass pic to PhotoImage function and set it to variable
    #pass background pic to background_pic function
    background_pic(bg)


    

    #sets textbox features
    e=Entry(root,width=50,bg="light gray",fg="black",borderwidth=5)
    e.pack()
    #insert text into textbox
    e.insert(0,"Enter Work ID#: ",)

    #makes texted in box delete by calling the main temp_text function
    e.bind("<FocusIn>",temp_text)

    
    #makes button return userClick function
    myButton= Button(root,text="Clock In", command=clockIn,padx=(10),pady=(10),background = "green", fg = "white")
    myButton.pack(padx=(10),pady=(10))
    

    myButton2= Button(root,text="Clock Out", command=clockOut,padx=(10),pady=(10),background = "red", fg = "white")

    myButton2.pack(padx=(10),pady=(10))

    myButton3= Button(root,text="Lunch Break", command=lunchIn,padx=(10),pady=(10),background = "green", fg = "white")

    myButton3.pack(padx=(10),pady=(10))
    
    myButton4= Button(root,text="Back from Lunch Break", command=lunchOut,padx=(10),pady=(10),background = "red", fg = "white")

    myButton4.pack(padx=(10),pady=(10))

    root.mainloop()
#(main line)-END     